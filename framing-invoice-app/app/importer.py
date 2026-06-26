"""
Import the `Baseline Price List` sheet of the framing materials workbook into
the baseline_items table.

The workbook is the SOURCE OF TRUTH. Re-importing replaces the baseline rows so
the database always reflects the latest spreadsheet. Item numbers are stored as
TEXT and a normalized_description is computed for fallback matching.

Column headers expected on the `Baseline Price List` sheet:
    Item #, Description, UM, Category, Baseline Unit Price, Baseline Rule,
    Lowest Price Seen, Highest Price Seen, Price Spread, Last Seen Price,
    Last Seen Invoice, Last Seen Date, Times Purchased, Total Qty, Total Paid,
    Notes
Only the fields the app needs are read; extras are ignored.
"""

import datetime

import openpyxl

from .db import connect
from .normalize import normalize_item_number, normalize_description

BASELINE_SHEET = "Baseline Price List"

# Map our DB columns -> the spreadsheet header text (case-insensitive match).
HEADER_MAP = {
    "item_number": "Item #",
    "description": "Description",
    "unit_measure": "UM",
    "category": "Category",
    "baseline_unit_price": "Baseline Unit Price",
    "lowest_price_seen": "Lowest Price Seen",
    "highest_price_seen": "Highest Price Seen",
    "last_seen_price": "Last Seen Price",
    "last_seen_invoice": "Last Seen Invoice",
    "last_seen_date": "Last Seen Date",
    "times_purchased": "Times Purchased",
    "total_qty": "Total Qty",
    "total_paid": "Total Paid",
}


def _coerce(value):
    """Make a cell value JSON/SQLite friendly (dates -> ISO strings)."""
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.date().isoformat() if isinstance(value, datetime.datetime) else value.isoformat()
    return value


def read_baseline_rows(xlsx_path: str):
    """Yield dict rows from the Baseline Price List sheet."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if BASELINE_SHEET not in wb.sheetnames:
        raise ValueError(
            f"Workbook is missing the '{BASELINE_SHEET}' sheet. "
            f"Found: {wb.sheetnames}"
        )
    ws = wb[BASELINE_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    # Build header text -> column index (case-insensitive).
    idx = {h.lower(): i for i, h in enumerate(header)}

    def col(field):
        want = HEADER_MAP[field].lower()
        return idx.get(want)

    for raw in rows[1:]:
        item_raw = raw[col("item_number")] if col("item_number") is not None else None
        # Skip blank / non-data rows.
        if item_raw is None or str(item_raw).strip() == "":
            continue
        rec = {}
        for field in HEADER_MAP:
            c = col(field)
            rec[field] = _coerce(raw[c]) if c is not None else None
        rec["item_number"] = normalize_item_number(item_raw)
        rec["normalized_description"] = normalize_description(rec.get("description"))
        yield rec


# Chart of accounts: header synonyms we accept (case-insensitive).
COA_HEADERS = {
    "account_number": ["account #", "account#", "account number", "account no",
                       "gl #", "gl account", "gl code", "code", "number", "acct #", "acct"],
    "account_name": ["account name", "name", "description", "account", "gl name"],
    "department": ["department", "dept", "division", "category group", "class", "phase"],
    "category": ["category", "type", "account type", "group"],
    "budget": ["budget", "budget amount", "amount", "target", "allowance"],
}


def read_coa_rows(path: str):
    """Yield {account_number, account_name, department, category} from a chart of
    accounts Excel or CSV. Header matching is forgiving so most exports work."""
    import csv as _csv
    ext = path.lower()
    if ext.endswith(".csv"):
        with open(path, newline="", encoding="utf-8", errors="replace") as f:
            rows = [list(r) for r in _csv.reader(f)]
    else:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    if not rows:
        return
    header = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    idx = {}
    for field, syns in COA_HEADERS.items():
        for i, h in enumerate(header):
            if h in syns:
                idx[field] = i
                break

    def get(row, field):
        i = idx.get(field)
        if i is None or i >= len(row) or row[i] is None:
            return None
        val = row[i]
        if isinstance(val, float) and val.is_integer():
            val = int(val)
        return str(val).strip() or None

    for raw in rows[1:]:
        num = get(raw, "account_number")
        name = get(raw, "account_name")
        if not num and not name:
            continue
        budget = get(raw, "budget")
        try:
            budget = float(str(budget).replace("$", "").replace(",", "")) if budget else None
        except (TypeError, ValueError):
            budget = None
        yield {"account_number": num, "account_name": name,
               "department": get(raw, "department"), "category": get(raw, "category"),
               "budget": budget}


def import_chart_of_accounts(path: str, db_path: str = None, replace: bool = True) -> int:
    """Import the chart of accounts into chart_of_accounts. Returns row count.

    Existing budgets are preserved on re-import (matched by account_number) so a
    re-upload of the account list doesn't wipe budgets typed in the app.
    """
    rows = list(read_coa_rows(path))
    with connect(db_path) as conn:
        prior = {r["account_number"]: r["budget"] for r in conn.execute(
            "SELECT account_number, budget FROM chart_of_accounts")} if replace else {}
        if replace:
            conn.execute("DELETE FROM chart_of_accounts")
        for rec in rows:
            if rec.get("budget") is None and rec["account_number"] in prior:
                rec["budget"] = prior[rec["account_number"]]
            conn.execute(
                """INSERT INTO chart_of_accounts
                   (account_number, account_name, department, category, budget)
                   VALUES (:account_number, :account_name, :department, :category, :budget)""",
                rec,
            )
    return len(rows)


# A price list attached to one account: flexible header matching.
PRICELIST_HEADERS = {
    "item_number": ["item #", "item#", "item no", "item number", "sku", "item", "code", "product #"],
    "description": ["description", "desc", "material", "item description", "product"],
    "unit_measure": ["um", "uom", "unit", "unit of measure", "u/m"],
    "baseline_unit_price": ["baseline unit price", "unit price", "price", "unit cost",
                            "each", "cost", "baseline price"],
    "category": ["category", "type"],
}


def import_price_list(path: str, gl_account: str, department: str = None,
                      db_path: str = None) -> int:
    """Import an itemized price list (Excel/CSV) and code every item to one
    account. Replaces just this account's items. Returns the count."""
    import csv as _csv
    if path.lower().endswith(".csv"):
        with open(path, newline="", encoding="utf-8", errors="replace") as f:
            rows = [list(r) for r in _csv.reader(f)]
    else:
        wb = openpyxl.load_workbook(path, data_only=True)
        # Prefer a "Baseline Price List" sheet if present (matches our format).
        sheet = next((wb[s] for s in wb.sheetnames if "baseline" in s.lower()
                      and "price" in s.lower()), wb[wb.sheetnames[0]])
        rows = [list(r) for r in sheet.iter_rows(values_only=True)]
    if not rows:
        return 0
    header = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    idx = {}
    for field, syns in PRICELIST_HEADERS.items():
        for i, h in enumerate(header):
            if h in syns:
                idx[field] = i
                break

    def get(row, field):
        i = idx.get(field)
        return row[i] if i is not None and i < len(row) else None

    def to_price(v):
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).replace("$", "").replace(",", "").strip()
        try:
            return float(s)
        except ValueError:
            return None

    recs = []
    for raw in rows[1:]:
        item = get(raw, "item_number")
        desc = get(raw, "description")
        price = to_price(get(raw, "baseline_unit_price"))
        if (item is None or str(item).strip() == "") and not desc:
            continue
        recs.append({
            "item_number": normalize_item_number(item),
            "description": str(desc).strip() if desc else None,
            "normalized_description": normalize_description(desc),
            "unit_measure": str(get(raw, "unit_measure")).strip() if get(raw, "unit_measure") else None,
            "category": str(get(raw, "category")).strip() if get(raw, "category") else "Material",
            "baseline_unit_price": price,
            "gl_account": gl_account,
            "department": department,
        })
    with connect(db_path) as conn:
        conn.execute("DELETE FROM baseline_items WHERE gl_account=?", (gl_account,))
        for rec in recs:
            conn.execute(
                """INSERT INTO baseline_items
                   (item_number, description, normalized_description, unit_measure,
                    category, baseline_unit_price, gl_account, department)
                   VALUES (:item_number, :description, :normalized_description,
                    :unit_measure, :category, :baseline_unit_price, :gl_account, :department)""",
                rec)
    return len(recs)


def import_baseline(xlsx_path: str, db_path: str = None, replace: bool = True) -> int:
    """Import the baseline sheet into baseline_items.

    With replace=True (default) the table is cleared first so the DB mirrors the
    spreadsheet exactly. Returns the number of rows imported.
    """
    rows = list(read_baseline_rows(xlsx_path))
    with connect(db_path) as conn:
        if replace:
            conn.execute("DELETE FROM baseline_items")
        for rec in rows:
            conn.execute(
                """INSERT INTO baseline_items
                   (item_number, description, normalized_description, unit_measure,
                    category, baseline_unit_price, lowest_price_seen,
                    highest_price_seen, last_seen_price, last_seen_invoice,
                    last_seen_date, times_purchased, total_qty, total_paid)
                   VALUES (:item_number, :description, :normalized_description,
                    :unit_measure, :category, :baseline_unit_price,
                    :lowest_price_seen, :highest_price_seen, :last_seen_price,
                    :last_seen_invoice, :last_seen_date, :times_purchased,
                    :total_qty, :total_paid)""",
                rec,
            )
    return len(rows)
