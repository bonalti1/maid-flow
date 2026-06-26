"""
Invoice / quote extraction -> structured JSON.

Supported uploads:
  * Spreadsheets (.xlsx, .xls)  -> parsed deterministically with openpyxl
  * CSV (.csv)                  -> parsed with the stdlib csv module
  * PDF (.pdf)                  -> text + tables via pdfplumber (optional dep)
  * Images (.png/.jpg/...)      -> OCR via pytesseract (optional dep)

The output is always the same structured dict:

    {
      "vendor_name": str | None,
      "invoice_number": str | None,
      "invoice_date": str | None,
      "property_or_job": str | None,
      "customer_po": str | None,
      "subtotal": float | None,
      "tax": float | None,
      "total": float | None,
      "extraction_status": "parsed" | "ocr" | "needs_review" | "error",
      "warnings": [str, ...],
      "line_items": [
         {quantity, unit_measure, item_number, description, unit_price,
          line_amount, category, extract_confidence}, ...
      ]
    }

OCR is inherently imperfect, so EVERY extracted result is meant to be edited by
the user before comparison. The comparison engine also independently downgrades
low-confidence lines to "OCR REVIEW".
"""

import csv
import io
import os
import re

import openpyxl

# Optional heavy deps — imported lazily so the app still runs without them.
try:
    import pdfplumber  # type: ignore
    HAVE_PDFPLUMBER = True
except Exception:  # pragma: no cover - depends on environment
    HAVE_PDFPLUMBER = False

try:
    import pytesseract  # type: ignore
    from PIL import Image  # type: ignore
    HAVE_OCR = True
except Exception:  # pragma: no cover - depends on environment
    HAVE_OCR = False


# Header synonyms we accept in spreadsheet/CSV invoices (lower-cased).
COLUMN_SYNONYMS = {
    "quantity": ["qty", "quantity", "qnty", "ordered"],
    "unit_measure": ["um", "uom", "unit", "unit of measure", "u/m"],
    "item_number": ["item #", "item#", "item no", "item number", "sku", "item", "product #", "code"],
    "description": ["description", "desc", "material", "item description", "product"],
    "unit_price": ["unit price", "price", "unit cost", "each", "price each", "u price"],
    "line_amount": ["line amount", "amount", "extended", "ext price", "total", "line total"],
    "category": ["category", "type"],
}


def _to_float(value):
    """Parse a price/quantity cell. Strips $ , and whitespace. None on failure."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("$", "").replace(",", "")
    if text in ("", "-"):
        return None
    # Handle parentheses negatives e.g. "(45.00)".
    neg = text.startswith("(") and text.endswith(")")
    text = text.strip("()")
    m = re.search(r"-?\d+(?:\.\d+)?", text)
    if not m:
        return None
    val = float(m.group())
    return -val if neg else val


def _match_columns(header):
    """Map column index -> our field name based on synonyms."""
    mapping = {}
    for i, h in enumerate(header):
        h_norm = str(h or "").strip().lower()
        for field, syns in COLUMN_SYNONYMS.items():
            if h_norm in syns and field not in mapping.values():
                mapping[i] = field
                break
    return mapping


def _rows_to_line_items(rows, confidence=1.0):
    """Turn header+data rows into line-item dicts using fuzzy header matching."""
    if not rows:
        return [], ["No rows found in upload."]
    # Find the header row: the first row that maps at least description+price.
    header_idx = None
    mapping = {}
    for ridx, row in enumerate(rows[:10]):
        m = _match_columns(row)
        fields = set(m.values())
        if "description" in fields and ("unit_price" in fields or "line_amount" in fields):
            header_idx, mapping = ridx, m
            break
    if header_idx is None:
        return [], ["Could not locate a header row with Description + Price columns."]

    warnings = []
    items = []
    for row in rows[header_idx + 1:]:
        rec = {"quantity": None, "unit_measure": None, "item_number": None,
               "description": None, "unit_price": None, "line_amount": None,
               "category": None, "extract_confidence": confidence}
        for cidx, field in mapping.items():
            val = row[cidx] if cidx < len(row) else None
            if field in ("quantity", "unit_price", "line_amount"):
                rec[field] = _to_float(val)
            else:
                rec[field] = str(val).strip() if val not in (None, "") else None
        # Skip fully empty rows.
        if not rec["description"] and rec["unit_price"] is None and not rec["item_number"]:
            continue
        items.append(rec)
    if not items:
        warnings.append("Header found but no data rows parsed.")
    return items, warnings


# --- metadata sniffing (vendor/invoice#/date/PO) from free text ------------

_DATE_RE = re.compile(r"\b(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}-\d{2}-\d{2})\b")
_INVOICE_RE = re.compile(r"(?:invoice|inv|quote|order)\s*#?\s*[:.]?\s*([A-Za-z0-9\-]+)", re.I)
_PO_RE = re.compile(r"(?:customer\s*po|po\s*#|p\.o\.)\s*[:.]?\s*([A-Za-z0-9\- ]+)", re.I)
_TOTAL_RE = re.compile(r"\btotal\b\s*[:.]?\s*\$?\s*([\d,]+\.\d{2})", re.I)
_TAX_RE = re.compile(r"(?:sales\s*)?tax\s*[:.]?\s*\$?\s*([\d,]+\.\d{2})", re.I)
_SUBTOTAL_RE = re.compile(r"sub\s*-?total\s*[:.]?\s*\$?\s*([\d,]+\.\d{2})", re.I)


def _sniff_metadata(text: str) -> dict:
    """Pull invoice header fields out of raw text with simple regexes."""
    meta = {"vendor_name": None, "invoice_number": None, "invoice_date": None,
            "property_or_job": None, "customer_po": None,
            "subtotal": None, "tax": None, "total": None}
    if not text:
        return meta
    m = _INVOICE_RE.search(text)
    if m:
        meta["invoice_number"] = m.group(1).strip()
    m = _DATE_RE.search(text)
    if m:
        meta["invoice_date"] = m.group(1)
    m = _PO_RE.search(text)
    if m:
        meta["customer_po"] = m.group(1).strip()
    m = _SUBTOTAL_RE.search(text)
    if m:
        meta["subtotal"] = _to_float(m.group(1))
    m = _TAX_RE.search(text)
    if m:
        meta["tax"] = _to_float(m.group(1))
    m = _TOTAL_RE.search(text)
    if m:
        meta["total"] = _to_float(m.group(1))
    # Vendor: first non-empty line is a decent heuristic.
    for line in text.splitlines():
        if line.strip():
            meta["vendor_name"] = line.strip()[:120]
            break
    return meta


# --- per-format parsers -----------------------------------------------------

def extract_xlsx(path: str) -> dict:
    """Parse an invoice/quote spreadsheet. Confidence is high (1.0)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    # Prefer an "Invoice Line Items"-style sheet if present, else first sheet.
    sheet = None
    for name in wb.sheetnames:
        if "line" in name.lower() and "item" in name.lower():
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb[wb.sheetnames[0]]
    rows = [list(r) for r in sheet.iter_rows(values_only=True)]
    items, warnings = _rows_to_line_items(rows, confidence=1.0)
    # Try to pull metadata from the parsed rows themselves.
    meta = _sniff_metadata("\n".join(
        " ".join(str(c) for c in r if c is not None) for r in rows[:5]
    ))
    meta.update(_metadata_from_columns(rows))
    result = {**meta, "line_items": items, "warnings": warnings,
              "extraction_status": "parsed"}
    return result


def _metadata_from_columns(rows):
    """If the sheet has explicit Invoice #/Date/PO columns, use them."""
    out = {}
    if not rows:
        return out
    header = [str(c or "").strip().lower() for c in rows[0]]

    def find(*names):
        for i, h in enumerate(header):
            if h in names:
                return i
        return None

    inv_i = find("invoice #", "invoice#", "invoice number")
    date_i = find("invoice date", "date")
    po_i = find("customer po", "po", "po #")
    job_i = find("job", "ship to / property", "property", "ship to")
    for r in rows[1:]:
        if inv_i is not None and inv_i < len(r) and r[inv_i] and "invoice_number" not in out:
            out["invoice_number"] = str(r[inv_i]).strip()
        if date_i is not None and date_i < len(r) and r[date_i] and "invoice_date" not in out:
            out["invoice_date"] = str(r[date_i])[:10]
        if po_i is not None and po_i < len(r) and r[po_i] and "customer_po" not in out:
            out["customer_po"] = str(r[po_i]).strip()
        if job_i is not None and job_i < len(r) and r[job_i] and "property_or_job" not in out:
            out["property_or_job"] = str(r[job_i]).strip()
        if {"invoice_number", "invoice_date"} <= out.keys():
            break
    return out


def extract_csv(path: str) -> dict:
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        rows = [list(r) for r in csv.reader(f)]
    items, warnings = _rows_to_line_items(rows, confidence=1.0)
    meta = _sniff_metadata("\n".join(",".join(map(str, r)) for r in rows[:5]))
    meta.update(_metadata_from_columns(rows))
    return {**meta, "line_items": items, "warnings": warnings,
            "extraction_status": "parsed"}


# Regex for a line-item-ish row in OCR/PDF text:
#   <item#> <description...> <qty> <unit_price> <line_amount>
# Very forgiving; the user edits afterward.
_LINE_RE = re.compile(
    r"^\s*(?P<item>[A-Za-z0-9\-]{2,12})?\s+(?P<desc>.+?)\s+"
    r"(?P<qty>\d+(?:\.\d+)?)\s+\$?(?P<price>\d+(?:\.\d+)?)\s+\$?(?P<amt>\d+(?:\.\d+)?)\s*$"
)


def _line_items_from_text(text: str, confidence: float):
    """Best-effort line-item parse from raw text (PDF/OCR)."""
    items, warnings = [], []
    for line in text.splitlines():
        m = _LINE_RE.match(line)
        if not m:
            continue
        qty = _to_float(m.group("qty"))
        price = _to_float(m.group("price"))
        amt = _to_float(m.group("amt"))
        # Plausibility filter: qty*price should be in the ballpark of amount.
        if qty and price and amt and abs(qty * price - amt) > max(1.0, 0.2 * amt):
            # Probably a mis-parse; keep it but lower confidence so it's flagged.
            line_conf = min(confidence, 0.45)
        else:
            line_conf = confidence
        items.append({
            "quantity": qty,
            "unit_measure": None,
            "item_number": (m.group("item") or "").strip() or None,
            "description": m.group("desc").strip(),
            "unit_price": price,
            "line_amount": amt,
            "category": None,
            "extract_confidence": line_conf,
        })
    if not items:
        warnings.append(
            "No line items could be auto-detected from the document text. "
            "Add them manually before comparing."
        )
    return items, warnings


def extract_pdf(path: str) -> dict:
    if not HAVE_PDFPLUMBER:
        return {**_sniff_metadata(""), "line_items": [],
                "warnings": ["pdfplumber not installed — cannot read PDFs. "
                             "Install it or upload a spreadsheet/CSV instead."],
                "extraction_status": "error"}
    text_parts = []
    table_items = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text_parts.append(page.extract_text() or "")
            for table in page.extract_tables() or []:
                its, _ = _rows_to_line_items(table, confidence=0.7)
                table_items.extend(its)
    text = "\n".join(text_parts)
    meta = _sniff_metadata(text)
    # Prefer structured tables; fall back to regex over the text.
    if table_items:
        items, warnings = table_items, []
    else:
        items, warnings = _line_items_from_text(text, confidence=0.7)
    status = "ocr" if items else "needs_review"
    return {**meta, "line_items": items, "warnings": warnings,
            "extraction_status": status}


def extract_image(path: str) -> dict:
    if not HAVE_OCR:
        return {**_sniff_metadata(""), "line_items": [],
                "warnings": ["pytesseract/Pillow not installed — cannot OCR images. "
                             "Install them or upload a spreadsheet/CSV instead."],
                "extraction_status": "error"}
    text = pytesseract.image_to_string(Image.open(path))
    meta = _sniff_metadata(text)
    items, warnings = _line_items_from_text(text, confidence=0.55)
    return {**meta, "line_items": items, "warnings": warnings,
            "extraction_status": "ocr" if items else "needs_review"}


IMAGE_EXTS = (".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".gif", ".webp")


def extract(path: str) -> dict:
    """Dispatch on file extension and return the structured JSON result.

    Spreadsheets/CSV are parsed deterministically. PDFs and images go through
    Claude vision (ai_extract) when it's configured — that's the reliable path
    for scanned/photographed invoices — and fall back to the local
    pdfplumber/tesseract parsers otherwise.
    """
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext in (".xlsx", ".xlsm", ".xls"):
            return extract_xlsx(path)
        if ext == ".csv":
            return extract_csv(path)
        if ext == ".pdf" or ext in IMAGE_EXTS:
            # Prefer AI vision for documents/photos when available.
            from . import ai_extract
            if ai_extract.ai_available():
                print(f"[extract] Using AI vision ({ai_extract.MODEL}) for {os.path.basename(path)}")
                result = ai_extract.extract_with_ai(path)
                # Only fall through to local parsing if AI hard-failed.
                if result.get("extraction_status") != "error":
                    print(f"[extract] AI read {len(result.get('line_items') or [])} line(s)")
                    return result
                ai_reason = "; ".join(result.get("warnings") or ["unknown error"])
                print(f"[extract] AI extraction failed -> falling back. Reason: {ai_reason}")
            else:
                # Explain WHY AI is off so the user can fix it.
                if not ai_extract.HAVE_ANTHROPIC:
                    ai_reason = "AI reading off: the 'anthropic' package isn't installed (run pip install -r requirements.txt)."
                else:
                    ai_reason = "AI reading off: no ANTHROPIC_API_KEY set in the terminal before starting the app."
                print(f"[extract] {ai_reason}")
            fallback = extract_pdf(path) if ext == ".pdf" else extract_image(path)
            # Surface the AI reason on the result so the UI can show it.
            fallback.setdefault("warnings", []).insert(0, ai_reason)
            return fallback
    except Exception as e:  # pragma: no cover - defensive
        return {**_sniff_metadata(""), "line_items": [],
                "warnings": [f"Extraction failed: {e}"],
                "extraction_status": "error"}
    return {**_sniff_metadata(""), "line_items": [],
            "warnings": [f"Unsupported file type: {ext}"],
            "extraction_status": "error"}
