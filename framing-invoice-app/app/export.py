"""
Export review results to Excel / CSV, and sync the baseline back to Excel.

* export_results_xlsx / export_results_csv -> the review table for one invoice
* export_baseline_xlsx -> write the current baseline_items back out to an .xlsx
  that mirrors the original `Baseline Price List` layout (optional sync).
"""

import csv
import io

import openpyxl

RESULT_COLUMNS = [
    ("item_number", "Item #"),
    ("description", "Description"),
    ("category", "Category"),
    ("quantity", "Qty"),
    ("unit_measure", "UM"),
    ("unit_price", "Invoice Unit Price"),
    ("baseline_unit_price", "Baseline Unit Price"),
    ("difference_per_unit", "Difference / Unit"),
    ("potential_overcharge", "Potential Overcharge"),
    ("status", "Status"),
    ("confidence_score", "Confidence"),
    ("notes", "Notes"),
]


def export_results_csv(rows) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([label for _, label in RESULT_COLUMNS])
    for r in rows:
        w.writerow([r.get(key) for key, _ in RESULT_COLUMNS])
    return buf.getvalue().encode("utf-8")


def export_results_xlsx(rows, summary=None) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Review"
    ws.append([label for _, label in RESULT_COLUMNS])
    for r in rows:
        ws.append([r.get(key) for key, _ in RESULT_COLUMNS])
    if summary:
        ws2 = wb.create_sheet("Summary")
        for k, v in summary.items():
            ws2.append([k, v])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


BASELINE_EXPORT_COLUMNS = [
    ("item_number", "Item #"),
    ("description", "Description"),
    ("unit_measure", "UM"),
    ("category", "Category"),
    ("baseline_unit_price", "Baseline Unit Price"),
    ("lowest_price_seen", "Lowest Price Seen"),
    ("highest_price_seen", "Highest Price Seen"),
    ("last_seen_price", "Last Seen Price"),
    ("last_seen_invoice", "Last Seen Invoice"),
    ("last_seen_date", "Last Seen Date"),
    ("times_purchased", "Times Purchased"),
    ("total_qty", "Total Qty"),
    ("total_paid", "Total Paid"),
]


def export_baseline_xlsx(baseline_rows) -> bytes:
    """Write the current baseline back to an .xlsx (Baseline Price List sheet)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Baseline Price List"
    ws.append([label for _, label in BASELINE_EXPORT_COLUMNS])
    for r in baseline_rows:
        ws.append([r.get(key) for key, _ in BASELINE_EXPORT_COLUMNS])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
